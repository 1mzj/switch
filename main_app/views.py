from django.shortcuts import render,redirect
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt, csrf_protect

from main_app import models
from django import forms
import json
from django.utils.timezone import datetime
from django.db.models import Count
from django.db.models.functions import TruncMonth, TruncYear

from main_app.utils.toolfunction import Pagination
from main_app.utils.update_ID import Update_ID
from .models import switch_datalist

#####模型预测###########
from main_app.Pre_model.tg_res_pre import predict_taoguan_res
from main_app.Pre_model.tg_tep_pre import predict_taoguan_tep
from main_app.Pre_model.dlq_res_pre import predict_duanluqi_res
from main_app.Pre_model.dlq_tep_pre import predict_duanluqi_tep
from main_app.Pre_model.glkg_res_pre import predict_gelikaiguan_res
from main_app.Pre_model.glkg_tep_pre import predict_gelikaiguan_tep
from main_app.Pre_model.hgq_res_pre import predict_huganqi_res
from main_app.Pre_model.hgq_tep_pre import predict_huganqi_tep
from main_app.Pre_model.jxb_res_pre import predict_jiexianban_res
from main_app.Pre_model.jxb_tep_pre import predict_jiexianban_tep
#####数据导出###########
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
import pytz
from datetime import datetime,timedelta
import urllib

###验证表单信息返回错误信息
from django.core.exceptions import ValidationError
####加密密码
import re
from main_app.utils.encrypt import md5
from django.contrib.auth.hashers import check_password
from django.core.cache import cache  # 用于记录登录失败次数

####生成图片验证码
from main_app.utils.create_code import check_code
from io import BytesIO
####退出登陆#######
from django.contrib.auth import logout

from django_ratelimit.decorators import ratelimit
from django.http import JsonResponse

@ratelimit(key='ip', rate='20/m', method='POST', block=True)
def my_view(request):
    return JsonResponse({'message': 'Success'})

# Create your views here.
class switch_Form(forms.ModelForm):
    class Meta:
        model = models.switch_datalist
        fields = '__all__'

    def __init__(self,*args,**kwargs):
        super().__init__(*args,**kwargs)

        for name,field in self.fields.items():
            field.widget.attrs = {"class":"form-control","placeholder":field.label}

class login_Form(forms.Form):

    user_name = forms.CharField(
        max_length=32,
        label = "用户名",
        widget = forms.TextInput(attrs={'placeholder': '请输入你的用户名'}),
        required=True,
    )

    user_password= forms.CharField(
        max_length=32,
        label = "密码",
        widget = forms.PasswordInput(attrs={'placeholder': '请输入你的密码'},render_value=True),
        required=True,
    )

    img_code = forms.CharField(
        max_length=32,
        label = "验证码",
        widget = forms.TextInput(attrs={'placeholder': '请输入图片验证码'}),
        required=True,
    )
    def clean_user_password(self):
        pwd = self.cleaned_data.get('user_password')
        return md5(pwd)

class user_Form(forms.ModelForm):
    ######在原有的数据库里面再加一个字段显示
    confirm_password = forms.CharField(
        label="确认密码",
        widget=forms.PasswordInput(attrs={'placeholder': '请再次输入您的密码'}),
    )
    class Meta:
        model = models.Userlist
        fields = ['user_name','user_phone','user_mail','user_password','confirm_password']
        widgets = {
            'user_password': forms.PasswordInput(),
            'user_name': forms.TextInput(attrs={'placeholder': '请输入你的用户名'}),
            'user_phone': forms.TextInput(attrs={'placeholder': '请输入您的手机号'}),
            'user_mail': forms.EmailInput(attrs={'placeholder': '请输入您的邮箱'}),
            'user_password': forms.PasswordInput(attrs={'placeholder': '请输入您的密码'}),
        }
    def __init__(self,*args,**kwargs):
        super().__init__(*args,**kwargs)
        self.error_messages = []
        # 遍历字段设置class属性
        for name, field in self.fields.items():
            # 不覆盖原有的attrs，只是更新class
            field.widget.attrs.update({"class": "form-control"})
        # 确保用户名唯一
    def clean_user_name(self):
        user_name = self.cleaned_data.get('user_name')
        if models.Userlist.objects.filter(user_name=user_name).exists():
            raise ValidationError('用户名已存在，请设置其他用户名')
        return user_name
    
    def clean_user_password(self):
        pwd = self.cleaned_data.get('user_password')
        if not re.search(r'[A-Za-z]', pwd) or not re.search(r'\d', pwd) or not re.search(r'[!@#$%^&*(),.?":{}|<>]', pwd):
            self.error_messages.append('密码必须包含字母、数字和特殊字符')
            raise ValidationError('密码必须包含字母、数字和特殊字符')
        if len(pwd) < 8:
            self.error_messages.append('密码最少8位')
            raise ValidationError('密码最少8位')
        username = self.cleaned_data.get('user_name')
        if username and pwd.lower() == username.lower():
            self.error_messages.append('密码不能与用户名相同')
            raise ValidationError('密码不能与用户名相同')
        return md5(pwd)
    
######验证两次密码输入是否一致（可以修改验证表单的信息）
#####这里的函数名尽量与要验证的变量保持一致
    def clean_confirm_password(self):
        # print(self.cleaned_data)  ###获得表的信息
        pwd = self.cleaned_data.get('user_password')
        if not pwd:
            print(self.error_messages)
            raise ValidationError(self.error_messages)
        confirm = md5(self.cleaned_data.get('confirm_password')) ###获得表单的确认密码输入信息
        if confirm != pwd:
            # print('两次输入的密码不一致')
            raise ValidationError('两次输入的密码不一致')
        return confirm    ####返回什么以后数据库字段就保存什么


screen_height = None
manager_table_display_num = None
def get_client_ip(request):
    """ 获取用户的 IP 地址 """
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip
def logout_view(request):
    # 销毁会话数据
    logout(request)
    return redirect('/login/')

def login(request):

    if request.method == 'GET':
        form = login_Form()
        return render(request, 'login.html', {'login_info': form})
        
    user_ip = get_client_ip(request)  # 获取用户 IP，防止不同用户影响
    cache_key = f'login_attempts_{user_ip}'  # 记录登录失败次数的 key
    lock_key = f'login_lock_{user_ip}'  # 记录锁定时间的 key
    lock_time = cache.get(lock_key)
    form = login_Form(data=request.POST)
    if lock_time:
        form.add_error('user_password', f'登录失败次数过多，请冷却期后再试！')
        return render(request, 'login.html', {'login_info': form})
    if form.is_valid():
        # print(form.cleaned_data)   ###（表单信息）
        user_input_code = form.cleaned_data.pop('img_code')
        img_code = request.session.get('img_code','')

        if img_code.upper() != user_input_code.upper():
            form.add_error('img_code',"验证码输入错误")
            return render(request, 'login.html', {'login_info': form})
        
        user_obj = models.Userlist.objects.filter(**form.cleaned_data).first()
        if not user_obj:
            # 登录失败，增加失败次数
            attempts = cache.get(cache_key, 0) + 1
            cache.set(cache_key, attempts, timeout=300)  # 失败次数保存 5 分钟
            remaining_attempts = 5 - attempts  # 计算剩余尝试次数

            if attempts >= 5:
                lock_time = datetime.now() + timedelta(minutes=5)  # 锁定 5 分钟
                cache.set(lock_key, lock_time.strftime("%H:%M:%S"), timeout=300)  # 记录解锁时间
                form.add_error('user_password', f'登录失败次数过多，请{lock_time.strftime("%H:%M:%S")} 后再试！')
                return render(request, 'login.html', {'login_info': form})

            form.add_error('user_password', f"用户名或密码错误，累计五次将限制。剩余{remaining_attempts}次尝试。")
            return render(request, 'login.html', {'login_info': form})

        # 登录成功，清除失败记录
        cache.delete(cache_key)
        cache.delete(lock_key)
        
        ####网站的session和cookie
        request.session['info'] = {'id':user_obj.id,'name': user_obj.user_name,'phone':user_obj.user_phone}
        request.session.set_expiry(60*60*3)  #session保存3小时

        return redirect("/index/")
    
    return render(request, 'login.html', {'login_info': form})
def create_code(request):
    img, code_string = check_code()

    ####记录验证码后续校验
    request.session['img_code'] = code_string
    request.session.set_expiry(60)  ###60s超时

    ###写入内存中并且获取返回前端
    stream = BytesIO()
    img.save(stream, 'png')
    return HttpResponse(stream.getvalue())

def index(request):
    ####检查密码是否是默认密码以及密码更改时间多久了
    password_warning = None
    default_password_warning = None
    DEFAULT_PASSWORD = "123456"  # 你的默认密码
    DEFAULT_PASSWORD_HASH = md5(DEFAULT_PASSWORD)  # 计算默认密码的MD5值
    user_id = request.session.get('info', {}).get('id')
    if user_id:
        user_obj = models.Userlist.objects.filter(id=user_id).first()
        # 计算是否超过半年
        if user_obj and user_obj.last_pwd_changed:
            half_year_ago = datetime.now() - timedelta(days=90)
            if user_obj.last_pwd_changed < half_year_ago:
                password_warning = True
        # 密码是否默认
        if user_obj.user_name.lower() != "admin":
            if user_obj.user_password == DEFAULT_PASSWORD_HASH:
                default_password_warning = True
    print(password_warning)
    print('default_password_warning:',default_password_warning)
    queryset = models.switch_datalist.objects.all()
    page_object =Pagination(request, queryset, page_size=4, page_displayNum=5)
    page_queryset = page_object.page_queryset
    context = {
        "all_data":queryset,
        "list_data": page_queryset,
        'password_warning': password_warning,
        'default_password_warning': default_password_warning,
    }
    if request.method == 'POST':
        pass
    return render(request, 'index.html',context)

@csrf_protect
def submit_json_data(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            ####数据读取########
            selected_dev_value = data.get('selected_dev_Value')
            pre_model = data.get('pre_model')
            input_data = [
                data.get('input1'),  # A相电流
                data.get('input2'),  # B相电流
                data.get('input3'),  # C相电流
                data.get('input4'),  # 转动角度
                data.get('input5'),  # 主轴扭矩 
            ]
            print("Input: ",selected_dev_value)
            print("Input_model ",pre_model)
            if pre_model == 'TEMP':
                print(input_data)
                if selected_dev_value == 'AAA':
                    ####选择隔离开关
                    tep = predict_gelikaiguan_tep(input_data)
                # elif selected_dev_value == 'BBB':
                #     ###选择断路器
                #     tep = predict_duanluqi_tep(input_data)
                    
                # elif selected_dev_value == 'CCC':
                #     ####选择电流互感器
                #     tep = predict_huganqi_tep(input_data)
                # elif selected_dev_value == 'DDD':
                #     ####选择套管
                #     tep = predict_taoguan_tep(input_data)
                # elif selected_dev_value == 'OOO':
                #     ####选择接线板
                #     tep = predict_jiexianban_tep(input_data)
                pre_temperature = round(float(tep), 2)
                context = {
                    'predicted_temperature': pre_temperature,
                }
            elif pre_model == 'RES':
                print(input_data)
                if selected_dev_value == 'AAA':
                    ####选择隔离开关
                    res = predict_gelikaiguan_res(input_data)
                elif selected_dev_value == 'BBB':
                    ###选择断路器
                    res = predict_duanluqi_res(input_data)
                elif selected_dev_value == 'CCC':
                    ####选择电流互感器
                    res = predict_huganqi_res(input_data)
                elif selected_dev_value == 'DDD':
                    ####选择套管
                    res = predict_taoguan_res(input_data)
                elif selected_dev_value == 'OOO':
                    ####选择接线板
                    res = predict_jiexianban_res(input_data)
                
                pre_resistance = round(float(res), 2)
                context = {
                    'resistance_value': pre_resistance,
                }
            
            return JsonResponse({'status': 'success', 'data': context})
            # Process data here
        except json.JSONDecodeError:
            print("error")
            return JsonResponse({'status': 'error', 'message': '无效的 JSON 数据'}, status=400)
    else:
        return JsonResponse({'status': 'error', 'message': '只接受 POST 请求'}, status=405)

def dataWeb(request):
    list_switch_Form = switch_Form()

    ###定义搜索内容字典####
    search_value = request.GET.get('condition')
    if search_value=="隔离开关":
        search_value = 1
    else:
        search_value = None
    # print(search_value)
    searchdata_dict={}
    if search_value:
        searchdata_dict = {"device_type__contains":search_value}
    
    # 获取数据并进行分页
    queryset = models.switch_datalist.objects.filter(**searchdata_dict)
    page_object =Pagination(request, queryset, page_size=16, page_displayNum=5)
    page_queryset = page_object.page_queryset
    page_string, no_more_content =page_object.html()
    

    # 如果记录数少于 8 条，填充空数据
    count = page_queryset.count()
    if count < 16:
        # 计算需要补充的条数
        missing_items = 16 - count
        
        # 创建一个空的数据列表
        empty_items = [None] * missing_items  # 或者使用空字典 [{}]，视你的需求而定
        
        # 合并原来的数据和空数据
        page_queryset = list(page_queryset) + empty_items

    context = {
        "list_switch_Form": list_switch_Form,
        "list":page_queryset,
        "page_string" :page_string,
    }

    return render(request, 'dataWeb.html', context)

#####将要编辑的数据返回到前端里面###########
Edit_obj_id = None
@csrf_protect
def operate_mes(request):
    global Edit_obj_id
    if request.method == 'POST':
        # 获取传递过来的 id 和操作类型
        obj_id = request.POST.get('id')
        action = request.POST.get('action')
        Edit_obj_id = obj_id
        try:
            # 获取对应的对象
            obj = models.switch_datalist.objects.get(id=obj_id)
            
            if action == 'edit':
                # # 编辑操作（例如修改设备类型）

                context ={
                    'device_type': obj.device_type,
                    'currentA': obj.currentA,
                    'currentB': obj.currentB,
                    'currentC': obj.currentC,
                    'angle': obj.angle,
                    'torque': obj.torque,
                    'add_time': obj.add_time.strftime('%Y-%m-%d %H:%M'),  # 格式化时间
                    'errors': {},  # 如果有验证错误可以返回错误信息
                }
                return JsonResponse({'status': 'success', 'data': context})
            elif action == 'delete':
                # 删除操作
                obj.delete()

                #记录操作
                user_info = request.session.get('info', {})
                print(user_info)
                user_name = user_info.get('name')
                user_phone = user_info.get('phone')
                models.Sys_log.objects.create(
                    event_type=3,  
                    user_name = user_name,
                    user_phone= user_phone,
                    record_time=datetime.now()
                )

                list_update = Update_ID(models.switch_datalist)

                return JsonResponse({'status': 'success', 'message': '删除成功！'})
            else:
                return JsonResponse({'status': 'error', 'message': '无效的操作！'})
        
        except models.switch_datalist.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': '数据不存在！'})
######获取编辑后的数据################

@csrf_protect
def save_edit(request):
    global Edit_obj_id
    if request.method == 'POST':
        device_type = request.POST.get('device_type')
        currentA = request.POST.get('currentA')
        currentB = request.POST.get('currentB')
        currentC = request.POST.get('currentC')
        angle = request.POST.get('angle')
        torque = request.POST.get('torque')
        add_time = request.POST.get('add_time')
        try:
            # 获取对象
            obj = models.switch_datalist.objects.get(id=Edit_obj_id)

            # 检查字段是否有变化，如果变化则更新字段
            if obj.device_type != device_type:
                obj.device_type = device_type  # 更新设备类型
            if obj.currentA != currentA:
                obj.currentA = currentA  # 更新A相电流
            if obj.currentB != currentB:
                obj.currentB = currentB  # 更新B相电流
            if obj.currentC != currentC:
                obj.currentC = currentC  # 更新C相电流
            if obj.angle != angle:
                obj.angle = angle  # 更新开关转动角度
            if obj.torque != torque:
                obj.torque = torque  # 更新开关主轴扭矩
            if obj.add_time.strftime('%Y-%m-%d %H:%M') != add_time:  # 注意：add_time是字符串，所以要与数据库中的格式对比
                # 将前端传来的时间转换成datetime类型
                obj.add_time = datetime.strptime(add_time, '%Y-%m-%d %H:%M')

            # 保存更新后的数据
            obj.save()
            
            Edit_obj_id = None

            #记录操作
            user_info = request.session.get('info', {})
            user_name = user_info.get('name')
            user_phone = user_info.get('phone')
            models.Sys_log.objects.create(
                event_type=2,  
                user_name = user_name,
                user_phone= user_phone,
                record_time=datetime.now()
            )

            return JsonResponse({'status': 'success', 'message': '数据已更新'})

        except models.switch_datalist.DoesNotExist:
            # 如果没有找到该对象
            return JsonResponse({'status': 'error', 'message': '数据不存在'})
        
    else:
        return JsonResponse({'status': 'error', 'message': '无效的请求'})
@csrf_protect
def submit_new_data(request):
    form = switch_Form(data=request.POST, files=request.FILES)
    if form.is_valid():
        device_type = request.POST.get('device_type')  # 获取原始的设备类型字段值
        add_time = request.POST.get('add_time')  # 获取原始的负荷电流字段值
        # print(device_type)
        # print(add_time)
        instance = form.save()

        #记录操作
        user_info = request.session.get('info', {})
        print(user_info)
        user_name = user_info.get('name')
        user_phone = user_info.get('phone')
        models.Sys_log.objects.create(
            event_type=1,  
            user_name = user_name,
            user_phone= user_phone,
            record_time=datetime.now()
        )

        list_update = Update_ID(models.switch_datalist)
        return JsonResponse({"status": True})

    return JsonResponse({"status":False,"error":form.errors})

def month(request):
    device_type_mapping = {
        1: "隔离开关",
    }
    # 统计每月不同设备类型的数据条数
    # 步骤 1: 按月份截取时间
    data = models.switch_datalist.objects.annotate(month=TruncMonth('add_time'))

    # 步骤 2: 按月和设备类型分组
    data = data.values('month', 'device_type')

    # 步骤 3: 统计每种设备类型每个月的条数
    data = data.annotate(count=Count('id'))

    # 步骤 4: 按月和设备类型排序
    data = data.order_by('month', 'device_type')
    # 输出查询结果
    result = []
    for entry in data:
        # 获取对应的中文设备类型名称
        device_type_name = device_type_mapping.get(entry['device_type'], '未知设备类型')

        # 以 "YYYY-MM" 格式输出月份
        month_str = entry['month'].strftime('%Y-%m')  # 格式化成 "YYYY-MM"
        
        result.append({
            'month': month_str,
            'device_type': device_type_name,
            'count': entry['count'],
        })

    count = len(result) 
    if count < 17:
        # 计算需要补充的条数
        missing_items = 17 - count
        
        # 创建一个空的数据列表
        empty_items = [None] * missing_items  # 或者使用空字典 [{}]，视你的需求而定
        
        # 合并原来的数据和空数据
        result = list(result) + empty_items
    # print(result)
    context = {
        'result': result,
    }
    return render(request, 'month.html', context)

def year(request):
    device_type_mapping = {
        1: "隔离开关",
    }

    # 步骤 1: 按年份截取时间
    data = models.switch_datalist.objects.annotate(year=TruncYear('add_time'))

    # 步骤 2: 按年和设备类型分组
    data = data.values('year', 'device_type')

    # 步骤 3: 统计每种设备类型每年的条数
    data = data.annotate(count=Count('id'))

    # 步骤 4: 按年和设备类型排序
    data = data.order_by('year', 'device_type')

    # 输出查询结果
    result = []
    for entry in data:
        # 获取对应的中文设备类型名称
        device_type_name = device_type_mapping.get(entry['device_type'], '未知设备类型')

        # 以 "YYYY" 格式输出年份
        year_str = entry['year'].strftime('%Y')  # 格式化成 "YYYY"
        
        result.append({
            'year': year_str,
            'device_type': device_type_name,
            'count': entry['count'],
        })

    # 计算数据条数
    count = len(result)
    if count < 17:
        # 计算需要补充的条数
        missing_items = 17 - count
        
        # 创建一个空的数据列表
        empty_items = [None] * missing_items  # 或者使用空字典 [{}]，视你的需求而定
        
        # 合并原来的数据和空数据
        result = list(result) + empty_items

    # 将数据传递到前端
    context = {
        'result': result,
    }
    return render(request, 'year.html', context)

def excel_export(request):
    # 获取 POST 请求中的参数
    year = request.GET.get('year')
    device_type = request.GET.get('device_type')
    device_type_inv_mapping = {
        "隔离开关": 1,
    }

    print(device_type)
    device_type = device_type_inv_mapping.get(device_type, None)
    print(device_type)
    # 筛选数据：按年份和设备类型过滤
    queryset = models.switch_datalist.objects.filter(add_time__year=year, device_type=device_type)

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "电力隔离开关运行数据"
    device_type_mapping = {
        1: "隔离开关",
    }
    # 写入表头
    headers = ["序号", "检测时间","设备类型", "A相电流(mA)", "B相电流(mA)", "C相电流(mA)", "转动角度(度)", "主轴扭矩(N·m)" ]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # 写入数据
    china_tz = pytz.timezone('Asia/Shanghai')
    row_idx = 2

    for obj in queryset:
        data_id = obj.id
        data_device_type = device_type_mapping.get(obj.device_type, "未知设备")
        data_currentA = obj.currentA  # A相电流
        data_currentB = obj.currentB  # B相电流
        data_currentC = obj.currentC  # C相电流
        data_angle = obj.angle  # 转动角度
        data_torque = obj.torque  #主轴扭矩
        data_add_time = obj.add_time  # 检测时间

        # 格式化检测时间
        if data_add_time is not None:
            data_add_time_native = data_add_time.astimezone(china_tz).replace(tzinfo=None)

        # 写入数据
        ws.cell(row=row_idx, column=1, value=row_idx-1)  # 序号按顺序生成
        ws.cell(row=row_idx, column=2, value=data_add_time_native if data_add_time is not None else data_add_time)  # 检测时间
        ws.cell(row=row_idx, column=3, value=data_device_type)  # 设备类型
        ws.cell(row=row_idx, column=4, value=data_currentA)  # A相电流
        ws.cell(row=row_idx, column=5, value=data_currentB)  # B相电流
        ws.cell(row=row_idx, column=6, value=data_currentC)  # C相电流
        ws.cell(row=row_idx, column=7, value=data_angle)  # 转动角度
        ws.cell(row=row_idx, column=8, value=data_torque)  # 主轴扭矩
        
        

        row_idx += 1  # 行号递增
    # 第一行加粗
    bold_font = Font(bold=True)
    for cell in ws[1]:  # ws[1] 表示第一行
        cell.font = bold_font  # 设置单元格的字体为加粗

    # 调整表格格式
    ws.row_dimensions[1].height = 20  # 第一行行高
    for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, max_col=ws.max_column), 1):  
        max_length = 0  
        for cell in col:
            # 靠左对齐
            cell.alignment = Alignment(horizontal='left')
            try:  
                if cell.value is not None:  
                    # 获取字符串长度（或其他适当的度量标准）  
                    length = len(str(cell.value))  
                    if length > max_length:  
                        max_length = length  
            except:  
                pass  

        # 根据最大长度计算列宽（这里只是一个简单的示例，你可能需要调整这个公式）  
        adjusted_width = (max_length + 2) * 1.5  

        # 设置列宽  
        column_letter = get_column_letter(col_idx)  
        ws.column_dimensions[column_letter].width = adjusted_width

    # 将工作簿保存到BytesIO对象中  
    output = BytesIO()  
    wb.save(output)

    # 重置BytesIO的读取位置到开始处  
    output.seek(0)
    current_date = datetime.now().strftime('%Y%m%d')
    device_type = device_type_mapping.get(device_type, None)
    yourname = f'电力隔离开关运行数据_{year}_{device_type}'  # 设备类型和年份加入文件名
    
    # 对文件名进行 URL 编码
    encoded_filename = urllib.parse.quote(f'{yourname}.xlsx')
    # 创建HTTP响应并发送Excel文件  
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')  
    response['Content-Disposition'] = f'attachment; filename*=utf-8\'\'{encoded_filename}'

    return response


@csrf_protect
def register(request):
    if request.method == 'GET':
        form = user_Form()
        context = {
            'user_info': form,
        }
        return render(request, 'register.html', context)
    else:
        form = user_Form(data=request.POST, files=request.FILES)
        # print(form)
        if form.is_valid():
            user = form.save(commit=False)  # 暂不提交到数据库
            user.last_pwd_changed = datetime.now()   # 赋值当前时间
            form.save()
            list_update = Update_ID(models.Userlist)
            return JsonResponse({"status": True})
        return JsonResponse({"status": False, 'error': form.errors})


def manager(request):
    global manager_table_display_num
    list_update = Update_ID(models.Userlist)
    searchdata_dict={}
    # 获取数据并进行分页
    queryset = models.Userlist.objects.all()
    page_object =Pagination(request, queryset, page_size=manager_table_display_num, page_displayNum=5)
    page_queryset = page_object.page_queryset
    page_string, no_more_content =page_object.html()
    context = {
        'user_info': page_queryset,
        'page_string': page_string,
    }

    return render(request, 'manager.html', context)

def manager_log(request):
    global manager_table_display_num
    list_update = Update_ID(models.Userlist)
    user_list = models.Userlist.objects.all() 

    searchdata_dict={}
    # 获取数据并进行分页
    queryset = models.Sys_log.objects.filter(**searchdata_dict)
    page_object =Pagination(request, queryset, page_size=manager_table_display_num, page_displayNum=5)
    page_queryset = page_object.page_queryset
    page_string, no_more_content =page_object.html()
    context = {
        'user_info': user_list,
        'log_info': page_queryset,
        'page_string': page_string,
    }
    return render(request, 'manager_log.html', context)


#####将要编辑的数据返回到前端里面###########
Edit_user_id = None
@csrf_protect
def operate_user(request):
    global Edit_user_id
    if request.method == 'POST':
        # 获取传递过来的 id 和操作类型
        obj_id = request.POST.get('id')
        action = request.POST.get('action')
        print('operate_user_Edit_user_id:',Edit_user_id)
        Edit_user_id = obj_id
        # print(Edit_obj_id)
        try:
            # 获取对应的对象
            obj = models.Userlist.objects.get(id=obj_id)
            
            if action == 'edit':
                context ={
                    'user_name': obj.user_name,
                    'user_phone': obj.user_phone,
                    'user_mail': obj.user_mail,
                    'user_password': None,
                    'errors': {},  # 如果有验证错误可以返回错误信息
                }
                return JsonResponse({'status': 'success', 'data': context})
            elif action == 'delete':
                # 删除操作
                obj.delete()
                
                # 记录操作
                user_info = request.session.get('info', {})
                user_name = user_info.get('name')
                user_phone = user_info.get('phone')
                models.Sys_log.objects.create(
                    event_type=5,  
                    user_name = user_name,
                    user_phone= user_phone,
                    record_time=datetime.now()
                )

                list_update = Update_ID(models.Userlist)
                if user_name != "Mz_admin":
                    logout(request)
                    return redirect('/login/')
                return JsonResponse({'status': 'success', 'message': '删除成功！'})
            else:
                return JsonResponse({'status': 'error', 'message': '无效的操作！'})
        
        except models.Userlist.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': '数据不存在！'})


@csrf_protect
def save_edit_user(request):
    global Edit_user_id
    if request.method == 'POST':
        user_name = request.POST.get('user_name')
        user_phone = request.POST.get('user_phone')
        user_mail = request.POST.get('user_mail')
        user_password = request.POST.get('user_password')
        try:
            # 获取对象
            obj = models.Userlist.objects.get(id=Edit_user_id)
            print('save_edit_user_Edit_user_id:',Edit_user_id)
            old_password = obj.user_password
            new_password = md5(user_password)
            # 检查字段是否有变化，如果变化则更新字段
            if obj.user_name == 'Mz_admin':
                if user_name != 'Mz_admin':
                    # Edit_user_id = None
                    return JsonResponse({'status': 'error', 'message': '系统管理员名称无法修改'})
                if old_password != new_password:
                    return JsonResponse({'status': 'error', 'message': '系统管理员密码无法修改'})
                if old_password == new_password:
                    if obj.user_phone != user_phone:
                        obj.user_phone = user_phone  
                    if obj.user_mail != user_mail:
                        obj.user_mail = user_mail  
            else:
                if obj.user_name != user_name:
                    obj.user_name = user_name  # 更新
                if obj.user_phone != user_phone:
                    obj.user_phone = user_phone  
                if obj.user_mail != user_mail:
                    obj.user_mail = user_mail  
                if old_password != new_password:
                    obj.user_password = new_password  
                    obj.last_pwd_changed = datetime.now()   # 赋值当前时间
                else:
                    return JsonResponse({'status': 'error', 'message': '新旧密码不能相同'})
                # 保存更新后的数据
            obj.save()
            Edit_user_id = None

            # 记录操作
            user_info = request.session.get('info', {})
            user_name = user_info.get('name')
            user_phone = user_info.get('phone')
            models.Sys_log.objects.create(
                event_type=4,  
                user_name = user_name,
                user_phone= user_phone,
                record_time=datetime.now()
            )

            return JsonResponse({'status': 'success', 'message': '数据已更新'})

        except models.Userlist.DoesNotExist:
            # 如果没有找到该对象
            return JsonResponse({'status': 'error', 'message': '数据不存在'})
        
    else:
        return JsonResponse({'status': 'error', 'message': '无效的请求'})

@csrf_protect
def reset_password(request):
    if request.method == 'POST':
        user_id = request.POST.get('user_id')

        try:
            # 获取用户对象
            obj = models.Userlist.objects.get(id=user_id)
            print(obj.user_name)
            # 重置密码为 '123456'（加密逻辑按需求修改）
            moren_password = "123456"
            new_password = md5(moren_password)
            obj.user_password = new_password
            obj.save()

            user_info = request.session.get('info', {})
            user_name = user_info.get('name')
            user_phone = user_info.get('phone')
            models.Sys_log.objects.create(
                event_type=6,  
                user_name = user_name,
                user_phone= user_phone,
                record_time=datetime.now()
            )
            
            # 返回成功响应
            return JsonResponse({'success': True})
        except models.Userlist.DoesNotExist:
            return JsonResponse({'success': False, 'error': '用户不存在'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': '无效的请求方式'})

@csrf_exempt
def get_screen_height(request):
    global screen_height
    global manager_table_display_num
    if request.method == "POST":
        # 从请求中解析出屏幕高度
        data = json.loads(request.body)
        screen_height = data.get('height')
        manager_table_display_num = int((screen_height * 0.85 *0.95 - 40- 37) / 39.6)
        manager_table_display_num = manager_table_display_num - 1
        # 可以在数据库中保存、做日志记录等操作
        print(f"Received screen height: {screen_height}")
        print(manager_table_display_num)
        
        return JsonResponse({"status": "success", "message": "Height received."})
    else:
        return JsonResponse({"status": "error", "message": "Invalid request method."})
    


def get_char_data(request):
    # data = switch_datalist.objects.values('currentA', 'add_time').order_by('add_time')   #按时间排序
    data = switch_datalist.objects.values('currentA', 'add_time',"currentB","currentC","angle","torque")
    result = [{ 'date': item['add_time'].strftime('%Y-%m-%d'),'currentA': item['currentA'],'currentB': item['currentB'],
               'currentC': item['currentC'],'angle': item['angle'],'torque': item['torque']} for item in data]
    return JsonResponse(result, safe=False)
    
